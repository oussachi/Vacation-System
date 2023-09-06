import openpyxl as op

def open_sheet(sheetName):
    wb = op.load_workbook("oussama.xlsx")
    sheet = wb[sheetName]
    return sheet


def getRow(sheetName, rowNumber):
    rowValues = []
    sheet = open_sheet(sheetName)
    max_col = sheet.max_column
 
    for i in range(1, max_col + 1):
        cell_obj = sheet.cell(row = rowNumber, column = i)
        rowValues.append(cell_obj.value)
    return rowValues


def getColumn(sheetName, columnNumber):
    columnValues = []
    sheet = open_sheet(sheetName)
    max_row = sheet.max_column
 
    for i in range(2, max_row + 1):
        cell_obj = sheet.cell(row = i, column = columnNumber)
        columnValues.appendrowValues(cell_obj.value)
    return columnValues


def getRowNumberOfSearch(sheetName, columnNumber, term):
    sheet = open_sheet(sheetName)
    max_row = sheet.max_row

    for i in range(2, max_row + 1):
        cell_obj = sheet.cell(row=i, column=columnNumber)
        if(cell_obj.value == str(term)):
            return i
    return None

def searchByColumn(sheetName, columnNumber, searchedValue):
    result = []
    sheet = open_sheet(sheetName)
    max_row = sheet.max_row

    for i in range(2, max_row + 1):
        cell_obj = sheet.cell(row=i, column=columnNumber)
        if(cell_obj.value == str(searchedValue)):
            result.append(getRow(sheetName, i))
    return result



# ---------------------------------- Case specific functions ------------------------------- #

def getEmployeesMatriculesByGRHCode(code):
    matricules = []
    sheet = open_sheet("Salariés")
    max_row = sheet.max_row

    # Column 3 is the GRH column
    for i in range(2, max_row + 1):
        cell_obj = sheet.cell(row=i, column=3)
        if(cell_obj.value == str(code)):
            matricule = sheet.cell(row=i, column=1)
            matricules.append(matricule.value)
    return matricules

def getEmployeesMatriculesByManagerCode(code):
    matricules = []
    sheet = open_sheet("Salariés")
    max_row = sheet.max_row

    # Column 4 is the Manager column
    for i in range(2, max_row + 1):
        cell_obj = sheet.cell(row=i, column=4)
        if(cell_obj.value == str(code)):
            matricule = sheet.cell(row=i, column=1)
            matricules.append(matricule.value)
    return matricules


def getUserByMatricule(matricule):
    user_data = searchByColumn("Salariés", 1, matricule)[0]
    return user_data

def getGRHByMatricule(matricule):
    user_data = searchByColumn("GRH", 1, matricule)[0]
    return user_data

def getManagerByMatricule(matricule):
    user_data = searchByColumn("Managers", 1, matricule)[0]
    return user_data